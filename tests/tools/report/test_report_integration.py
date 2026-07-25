"""Integration tests for report tool adapters with REAL skill data.

These tests pipe real skill results through the report adapters (table, chart,
xlsx) to verify the full pipeline works end-to-end. They skip gracefully when
the required databases are not synced.

Run just these tests:
    python -m pytest tests/tools/report/test_report_integration.py -v -W error --tb=short
"""
from __future__ import annotations

from pathlib import Path

import pytest

TICKERS = ["PETR4", "KLBN11", "SUZB3"]


# ── Reuse the skip-when-missing fixtures ─────────────────────────────────────

@pytest.fixture
def cvm_dbs_available():
    """Skip if any CVM DB needed for the full pipeline is missing."""
    missing = []
    for db_name, connect_fn in [
        ("dfp", "data_sources.cvm._db.connect_dfp"),
        ("itr", "data_sources.cvm._db.connect_itr"),
        ("fre", "data_sources.cvm._db.connect_fre"),
        ("cad", "data_sources.cvm._db.connect_cad"),
    ]:
        try:
            mod_path, fn_name = connect_fn.rsplit(".", 1)
            import importlib
            mod = importlib.import_module(mod_path)
            conn = getattr(mod, fn_name)(read_only=True)
            conn.close()
        except (FileNotFoundError, ImportError):
            missing.append(db_name)
    if missing:
        pytest.skip(f"CVM DBs not synced: {missing}")
    return True


@pytest.fixture
def bridge_available():
    try:
        from data_sources.cvm.bridge.query_engine import lookup
        r = lookup(ticker="PETR4")
        if r.get("status") != "ok":
            pytest.skip("Bridge not synced")
    except (FileNotFoundError, ImportError):
        pytest.skip("Bridge not synced")
    return True


@pytest.fixture
def b3_dividends_available():
    try:
        from data_sources.b3.dividends.query_engine import dividends
        r = dividends(ticker="PETR4", limit=1)
        if r.get("status") == "not_synced":
            pytest.skip("B3 dividends not synced")
    except (FileNotFoundError, ImportError):
        pytest.skip("B3 dividends not synced")
    return True


@pytest.fixture
def cotahist_available():
    try:
        from data_sources.b3.cotahist.query_engine import query
        r = query(ticker="PETR4", limit=1, market_type=10)
        if r.get("status") == "not_synced":
            pytest.skip("COTAHIST not synced")
    except (FileNotFoundError, ImportError):
        pytest.skip("COTAHIST not synced")
    return True


# ── Table action integration tests ───────────────────────────────────────────

class TestTableAdapters:
    """report(action="table") with real skill data through adapters."""

    def test_table_financials_quarterly(self, cvm_dbs_available, bridge_available, tmp_path, monkeypatch):
        from skills.cvm.financials.financials import quarterly
        from tools.report_ops import table
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = quarterly(company="PETR4", periods=4)
        r = table.build("int-fin", "PETR4 Financials", data,
                        {"adapter": "financials_quarterly", "theme": "dark"})
        assert r["type"] == "table"
        assert r["sections"] > 0
        assert Path(r["html_path"]).exists()

    def test_table_valuation_unit_ticker(self, cvm_dbs_available, bridge_available, tmp_path, monkeypatch):
        """Verify valuation adapter works for UNIT tickers (KLBN11)."""
        from skills.cvm.valuation.valuation import ratios
        from tools.report_ops import table
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = ratios(company="KLBN11")
        r = table.build("int-val-unit", "KLBN11 Valuation", data,
                        {"adapter": "valuation_ratios", "theme": "dark"})
        assert r["type"] == "table"
        assert Path(r["html_path"]).exists()
        # Verify the HTML contains the P/L value (not None)
        html = Path(r["html_path"]).read_text(encoding="utf-8")
        ratios_data = data.get("ratios", {})
        pl = ratios_data.get("p_l")
        if pl is not None:
            # P/L should appear in the rendered table (formatted as "num" spec)
            assert "P/L" in html

    def test_table_comparison_side_by_side(self, cvm_dbs_available, bridge_available,
                                            b3_dividends_available, tmp_path, monkeypatch):
        from skills.cvm.comparison.comparison import side_by_side
        from tools.report_ops import table
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = side_by_side(tickers=TICKERS)
        r = table.build("int-cmp", " vs ".join(TICKERS), data,
                        {"adapter": "comparison_side_by_side", "theme": "dark"})
        assert r["type"] == "table"
        assert r["sections"] == 3  # valuation + financials + dividends
        assert Path(r["html_path"]).exists()


# ── Chart action integration tests ───────────────────────────────────────────

class TestChartAdapters:
    """report(action="chart") with real skill data through adapters."""

    def test_chart_financials_trend(self, cvm_dbs_available, bridge_available, tmp_path, monkeypatch):
        from skills.cvm.financials.financials import quarterly
        from tools.report_ops import charts
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = quarterly(company="PETR4", periods=8)
        r = charts.build("int-chart-fin", "PETR4 Trends", data,
                         {"chart_type": "line", "adapter": "financials_quarterly_chart", "theme": "dark"})
        assert r["type"] == "chart"
        assert r["chart_type"] == "line"
        assert Path(r["html_path"]).exists()

    def test_chart_cotahist_close(self, cotahist_available, tmp_path, monkeypatch):
        from data_sources.b3.cotahist.query_engine import query
        from tools.report_ops import charts
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = query(ticker="PETR4", limit=30, market_type=10)
        r = charts.build("int-chart-cot", "PETR4 Price", data,
                         {"chart_type": "line", "adapter": "cotahist_close_chart", "theme": "dark"})
        assert r["type"] == "chart"
        assert Path(r["html_path"]).exists()


# ── xlsx export integration tests ────────────────────────────────────────────

class TestXlsxExport:
    """report(action="export", format="xlsx") with real skill data."""

    def test_xlsx_comparison(self, cvm_dbs_available, bridge_available,
                              b3_dividends_available, tmp_path, monkeypatch):
        from skills.cvm.comparison.comparison import side_by_side
        from tools.report_ops import export
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = side_by_side(tickers=TICKERS)
        r = export.run("int-xlsx", "Comparison", data,
                       {"format": "xlsx", "adapter": "comparison_side_by_side"})
        assert r["status"] == "success"
        assert r.get("xlsx_path")
        assert Path(r["xlsx_path"]).exists()
        assert r["sheets"] == 3  # valuation + financials + dividends
        # Verify xlsx has native numeric cells
        try:
            from openpyxl import load_workbook
            wb = load_workbook(r["xlsx_path"])
            assert len(wb.sheetnames) == 3
            # First sheet should have numeric cells (not all strings)
            ws = wb[wb.sheetnames[0]]
            numeric_found = False
            for row in ws.iter_rows(min_row=2, max_row=3, min_col=2, max_col=5):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        numeric_found = True
                        break
                if numeric_found:
                    break
            assert numeric_found, "xlsx should have native numeric cells"
        except ImportError:
            pass  # openpyxl not installed — skip deep verify

    def test_xlsx_financials(self, cvm_dbs_available, bridge_available, tmp_path, monkeypatch):
        from skills.cvm.financials.financials import annual
        from tools.report_ops import export
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = annual(company="PETR4", periods=3)
        r = export.run("int-xlsx-fin", "PETR4 Annual", data,
                       {"format": "xlsx", "adapter": "financials_annual"})
        assert r["status"] == "success"
        assert Path(r["xlsx_path"]).exists()
