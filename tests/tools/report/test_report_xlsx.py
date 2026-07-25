"""Tests for report_ops/export.py — xlsx export path.

The pdf/png path (Playwright) is already covered by test_report_export.py.
These tests focus on the v1.2 xlsx path: native numeric cells, per-column
Excel number formats, multi-sheet, adapter support, openpyxl-missing fallback.

openpyxl is a test dependency (it's in requirements.txt). If absent, the
graceful-warning branch is exercised.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

from tools.report_ops import export


# ── Fixtures ─────────────────────────────────────────────────────────────────

TABLE_DATA = {
    "company": "PETR4",
    "sections": [{
        "title": "Quarterly Summary",
        "columns": ["Período", "Receita", "EBITDA", "Marg. EBITDA", "ROE"],
        "rows": [
            ["2T2025", 120_000, 30_000, 0.25, 0.20],
            ["1T2025", 100_000, 25_000, 0.25, 0.18],
        ],
        "formats": {"Receita": "brl", "EBITDA": "brl",
                    "Marg. EBITDA": "pct", "ROE": "pct"},
    }],
}

FINANCIALS_SKILL = {
    "status": "ok", "company": "PETROBRAS", "period_type": "quarterly",
    "periods": [
        {"period": "1T2025",
         "metrics": {"receita_liquida": 100_000, "lucro_bruto": 30_000,
                     "ebit": 20_000, "ebitda": 25_000, "lucro_liquido": 15_000,
                     "fco": 22_000},
         "ratios": {"marg_bruta": 0.30, "marg_ebitda": 0.25,
                    "marg_liquida": 0.15, "roe": 0.18}},
    ],
}


# ── Coercion ─────────────────────────────────────────────────────────────────

class TestCoerceXlsxData:
    def test_dict_passthrough(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        out = export._coerce_xlsx_data(TABLE_DATA, {})
        assert out["sections"][0]["title"] == "Quarterly Summary"

    def test_list_of_sections_wrapped(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        out = export._coerce_xlsx_data(TABLE_DATA["sections"], {})
        assert "sections" in out

    def test_adapter_applied(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        out = export._coerce_xlsx_data(FINANCIALS_SKILL, {"adapter": "financials_quarterly"})
        assert out["sections"]  # adapter flattened it

    def test_json_file_path_loaded(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        # _coerce_xlsx_data uses resolve_path; conftest patches it permissively
        jf = tmp_path / "data.json"
        jf.write_text(json.dumps(TABLE_DATA), encoding="utf-8")
        out = export._coerce_xlsx_data(str(jf), {})
        assert out["sections"][0]["title"] == "Quarterly Summary"

    def test_no_sections_raises(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        with pytest.raises(ValueError, match="sections"):
            export._coerce_xlsx_data({"kpis": []}, {})

    def test_non_dict_non_str_raises(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        with pytest.raises(ValueError, match="dict"):
            export._coerce_xlsx_data(42, {})


# ── Sheet name sanitization ──────────────────────────────────────────────────

class TestSheetName:
    def test_strips_invalid_chars(self):
        assert export._safe_sheet_name("Receita: 2024", set()) == "Receita 2024"

    def test_truncates_to_31_chars(self):
        long = "A" * 50
        assert len(export._safe_sheet_name(long, set())) <= 31

    def test_dedup_with_suffix(self):
        used = {"sheet1"}
        assert export._safe_sheet_name("Sheet1", used) == "Sheet1 (1)"
        # second collision increments
        assert export._safe_sheet_name("Sheet1", used) == "Sheet1 (2)"

    def test_empty_falls_back_to_sheet(self):
        assert export._safe_sheet_name("", set()) == "Sheet"


# ── xlsx export end-to-end ───────────────────────────────────────────────────

class TestExportXlsx:
    def test_xlsx_from_table_data(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        r = export.run("xlsx-1", "Financials", TABLE_DATA, {"format": "xlsx"})
        assert r["status"] == "success"
        assert r["sheets"] == 1
        assert r["adapter"] == ""
        xlsx_path = Path(r["xlsx_path"])
        assert xlsx_path.exists()
        assert xlsx_path.suffix == ".xlsx"

    def test_xlsx_native_numeric_cells(self, tmp_path, monkeypatch):
        """Numeric cells must stay native (not strings) with Excel formats."""
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        r = export.run("xlsx-2", "F", TABLE_DATA, {"format": "xlsx"})
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not installed")
        wb = load_workbook(r["xlsx_path"])
        ws = wb[wb.sheetnames[0]]
        # header
        assert [c.value for c in ws[1]] == ["Período", "Receita", "EBITDA", "Marg. EBITDA", "ROE"]
        # row 2: ["2T2025", 120000, 30000, 0.25, 0.20]
        assert ws["A2"].value == "2T2025"
        # numeric cells native
        assert ws["B2"].value == 120_000
        assert isinstance(ws["B2"].value, (int, float))
        # Excel number format applied: BRL
        assert ws["B2"].number_format == '"R$ "#,##0.00'
        # pct: 0.25 with 0.00% format
        assert ws["D2"].value == 0.25
        assert ws["D2"].number_format == "0.00%"

    def test_xlsx_none_cells_empty(self, tmp_path, monkeypatch):
        """None values leave numeric cells empty (not 0, not 'None')."""
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = {"sections": [{
            "title": "T", "columns": ["A", "B"],
            "rows": [["x", None], ["y", 5]],
            "formats": {"B": "brl"},
        }]}
        r = export.run("xlsx-3", "None", data, {"format": "xlsx"})
        from openpyxl import load_workbook
        wb = load_workbook(r["xlsx_path"])
        ws = wb[wb.sheetnames[0]]
        assert ws["B2"].value is None  # None -> empty cell
        assert ws["B3"].value == 5

    def test_xlsx_multi_sheet(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = {"sections": [
            {"title": "Q1", "columns": ["X"], "rows": [[1]]},
            {"title": "Q2", "columns": ["X"], "rows": [[2]]},
            {"title": "Q3", "columns": ["X"], "rows": [[3]]},
        ]}
        r = export.run("xlsx-4", "Multi", data, {"format": "xlsx"})
        assert r["sheets"] == 3
        from openpyxl import load_workbook
        wb = load_workbook(r["xlsx_path"])
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames[0] == "Q1"

    def test_xlsx_with_adapter(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        r = export.run("xlsx-5", "PETR4", FINANCIALS_SKILL,
                       {"format": "xlsx", "adapter": "financials_quarterly"})
        assert r["status"] == "success"
        assert r["adapter"] == "financials_quarterly"
        assert r["sheets"] == 1
        from openpyxl import load_workbook
        wb = load_workbook(r["xlsx_path"])
        ws = wb[wb.sheetnames[0]]
        # header row should contain money columns
        header = [c.value for c in ws[1]]
        assert "Receita Líquida" in header
        assert "EBITDA" in header

    def test_xlsx_text_spec_cells_are_strings(self, tmp_path, monkeypatch):
        """text-spec columns stay as strings (not coerced to number)."""
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        data = {"sections": [{
            "title": "T", "columns": ["Ticker", "Preço"],
            "rows": [["PETR4", 38.5]],
            "formats": {"Ticker": "text", "Preço": "brl_full"},
        }]}
        r = export.run("xlsx-6", "Text", data, {"format": "xlsx"})
        from openpyxl import load_workbook
        wb = load_workbook(r["xlsx_path"])
        ws = wb[wb.sheetnames[0]]
        assert ws["A2"].value == "PETR4"
        assert ws["B2"].value == 38.5
        assert ws["B2"].number_format == '"R$ "#,##0.00'


# ── Format dispatch (pdf/png still work, xlsx is the new branch) ─────────────

class TestRunDispatch:
    def test_run_xlsx_branches_to_export_xlsx(self, tmp_path, monkeypatch):
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        r = export.run("d-1", "F", TABLE_DATA, {"format": "xlsx"})
        assert "xlsx_path" in r
        assert "pdf_path" not in r
        assert "png_path" not in r

    def test_run_pdf_branches_to_export_html(self, tmp_path, monkeypatch):
        """pdf format must dispatch to _export_html (Playwright path), NOT _export_xlsx.

        Robust to whether Playwright is installed: if installed, a pdf_path is
        produced; if absent, a graceful warning is returned. Either way the
        dispatch must NOT produce an xlsx_path.
        """
        monkeypatch.setattr("tools.report_ops.paths.cfg.workspace_root", tmp_path)
        # need a real HTML file for _export_html to get past the path check
        html = tmp_path / "report.html"
        html.write_text("<html></html>", encoding="utf-8")
        r = export.run("d-2", "R", str(html), {"format": "pdf"})
        assert r["status"] == "success"
        # Dispatch proof: html path returned, NO xlsx_path key
        assert "html_path" in r
        assert "xlsx_path" not in r
        # Either Playwright produced a pdf, or it's absent (warning set)
        assert "pdf_path" in r or "playwright not installed" in (r.get("warning") or "")
