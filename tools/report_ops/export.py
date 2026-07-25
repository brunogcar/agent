"""report_ops/export.py - PDF/PNG/xlsx export.

Two export paths share one action:

  PDF / PNG  — Playwright captures an existing HTML report (lazy import, optional).
  xlsx       — openpyxl writes table data (or a skill result via an adapter) to
               a multi-sheet workbook. Native numeric cells keep Excel number
               formats so they stay sortable/summable.

Both heavy libs (playwright, openpyxl) are imported lazily and degrade to a
graceful warning if absent.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from core.path_guard import resolve_path
from tools.report_ops.paths import report_out_dir


def run(
    trace_id: str,
    title: str,
    data: Any,
    config: dict,
) -> dict:
    """Export to PDF/PNG (from an HTML file) or xlsx (from table data / skill JSON).

    config["format"]: "pdf" | "png" | "xlsx"
      - pdf/png: data = path to an existing HTML report file.
      - xlsx:    data = table-shape dict, a list of sections, a skill result
                       (with config["adapter"]), or a path to a local .json file.
    """
    fmt = config.get("format", "pdf").lower()

    if fmt == "xlsx":
        return _export_xlsx(trace_id, title, data, config)

    return _export_html(trace_id, title, data, config, fmt)


# ── PDF / PNG via Playwright ─────────────────────────────────────────────────

def _export_html(trace_id: str, title: str, data: Any, config: dict, fmt: str) -> dict:
    html_path_str = data if isinstance(data, str) else config.get("html_path", "")
    if not html_path_str:
        raise ValueError("data must be the path to an existing HTML file")

    # v1.1 fix: default to workspace root since reports live in workspace/reports/
    p, err = resolve_path(html_path_str, default_root="workspace")
    if err:
        raise ValueError(err)
    if not p.exists():
        raise ValueError(f"HTML file not found: {p}")

    out_dir = report_out_dir(trace_id)
    safe_title = "".join(c if c.isalnum() or c in "-_" else "_" for c in (title or "export"))

    try:
        from playwright.sync_api import sync_playwright  # lazy
    except ImportError:
        return {
            "status": "success",
            "html_path": str(p),
            "pdf_path": None,
            "png_path": None,
            "warning": "playwright not installed - install with: pip install playwright",
        }

    export_path = out_dir / f"{safe_title}.{fmt}"

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f"file:///{p.resolve().as_posix()}")
        # Expand all tabs/collapsibles before export
        page.evaluate("""
            document.querySelectorAll('.tab-panel').forEach(el => el.classList.add('active'));
            document.querySelectorAll('.collapsible').forEach(el => el.classList.add('open'));
            document.querySelectorAll('.sidebar, .topbar, .btn-icon').forEach(el => el.style.display='none');
        """)
        if fmt == "pdf":
            page.pdf(path=str(export_path), format="A4", print_background=True)
        else:
            page.screenshot(path=str(export_path), full_page=True)
        browser.close()

    return {
        "status": "success",
        "html_path": str(p),
        f"{fmt}_path": str(export_path),
    }


# ── xlsx via openpyxl ────────────────────────────────────────────────────────

def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet names: <=31 chars, no :\\/?*[], unique."""
    clean = "".join(c for c in (name or "Sheet") if c not in ':\\/?*[]')
    clean = clean.strip()[:31] or "Sheet"
    base = clean
    i = 1
    while clean.lower() in used:
        suffix = f" ({i})"
        clean = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(clean.lower())
    return clean


def _coerce_xlsx_data(data: Any, config: dict) -> dict:
    """Resolve ``data`` into a table-shape dict for xlsx export.

    Accepts: a dict (table shape or skill result), a list of sections, or a
    path to a local .json file. Applies config["adapter"] when set.
    """
    adapter = (config.get("adapter") or "").strip()

    # Load from JSON file path
    if isinstance(data, str):
        p, err = resolve_path(data)
        if err:
            raise ValueError(err)
        if not p.exists():
            raise ValueError(f"File not found: {p}")
        data = json.loads(p.read_text(encoding="utf-8"))

    if adapter:
        from tools.report_ops.adapters import apply_adapter
        data = apply_adapter(adapter, data)

    if isinstance(data, list):
        data = {"sections": data}
    if not isinstance(data, dict):
        raise ValueError(
            "xlsx data must be a dict (table shape or skill result), a list of "
            "sections, or a .json file path."
        )
    if not data.get("sections"):
        raise ValueError("xlsx data has no 'sections' to export")
    return data


def _export_xlsx(trace_id: str, title: str, data: Any, config: dict) -> dict:
    from tools.report_ops.table import _normalize_section
    from tools.report_ops.formats import excel_format, is_numeric_spec, _is_missing

    table_data = _coerce_xlsx_data(data, config)
    sections = [_normalize_section(s, i) for i, s in enumerate(table_data.get("sections") or [])]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {
            "status": "success",
            "xlsx_path": None,
            "warning": "openpyxl not installed - install with: pip install openpyxl",
        }

    out_dir = report_out_dir(trace_id)
    safe_title = "".join(c if c.isalnum() or c in "-_" else "_" for c in (title or "export"))
    export_path = out_dir / f"{safe_title}.xlsx"

    wb = Workbook()
    # Remove the default sheet — we'll create named ones per section
    default_ws = wb.active
    wb.remove(default_ws)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    used_names: set[str] = set()

    for sec in sections:
        ws = wb.create_sheet(_safe_sheet_name(sec["title"], used_names))
        columns = sec["columns"]
        col_formats = sec["col_formats"]

        # Header row
        for c_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(col))
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"

        # Data rows — native numbers for numeric specs, strings otherwise
        for r_idx, row in enumerate(sec["rows"], start=2):
            for c_idx, (value, spec) in enumerate(zip(row, col_formats), start=1):
                if _is_missing(value):
                    # Leave numeric cells empty (so they don't count as 0); text -> ""
                    cell = ws.cell(row=r_idx, column=c_idx, value=None)
                    continue
                if is_numeric_spec(spec):
                    try:
                        num = float(value)
                        cell = ws.cell(row=r_idx, column=c_idx, value=num)
                        cell.number_format = excel_format(spec)
                        cell.alignment = Alignment(horizontal="right")
                        continue
                    except (TypeError, ValueError):
                        pass
                # text / fallback
                cell = ws.cell(row=r_idx, column=c_idx, value=str(value))

        # Auto-size columns (cap width for long text)
        for c_idx, col in enumerate(columns, start=1):
            max_len = len(str(col))
            for row in sec["rows"][:50]:  # sample first 50 rows for speed
                v = row[c_idx - 1] if c_idx - 1 < len(row) else ""
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = min(max_len + 2, 48)

    wb.save(str(export_path))

    return {
        "status": "success",
        "xlsx_path": str(export_path),
        "sheets": len(sections),
        "adapter": (config.get("adapter") or "").strip(),
    }
