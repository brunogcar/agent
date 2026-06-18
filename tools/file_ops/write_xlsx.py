"""
Write Excel action handler.
"""

from __future__ import annotations

import json as _json
from pathlib import Path

from tools.file_ops.helpers import _safe_resolve
from tools.file_ops._registry import register_action

@register_action("file", "write_xlsx")
def _handle_write_xlsx(path: str = "", content: str = "", **kwargs) -> dict:
    """Write data to an Excel file using pandas."""
    p, err = _safe_resolve(path)
    if err:
        return {"status": "error", "error": err}
    if not content and not isinstance(content, (dict, list)):
        return {"status": "error", "error": "content is required for write_xlsx"}
    if p.suffix.lower() not in (".xlsx", ".xls"):
        p = p.with_suffix(".xlsx")

    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Parse content if it's a string
        if isinstance(content, str):
            try:
                content = _json.loads(content)
            except Exception:
                return {"status": "error",
                        "error": "content string could not be parsed as JSON"}

        p.parent.mkdir(parents=True, exist_ok=True)

        sheets_written = []

        # Build dataframes first, then write in a tight scope
        dfs = []
        if isinstance(content, dict):
            content = {str(k): v for k, v in content.items()}
            first_val = next(iter(content.values()), None)

            if isinstance(first_val, list) and all(isinstance(v, list) for v in content.values()):
                for sheet_name, rows in content.items():
                    safe_name = str(sheet_name)[:31]
                    try:
                        if rows and isinstance(rows[0], dict):
                            rows = [{str(k): v for k, v in r.items()} for r in rows]
                            df = pd.DataFrame(rows)
                        elif rows and isinstance(rows[0], list):
                            df = pd.DataFrame(rows[1:], columns=rows[0])
                        else:
                            df = pd.DataFrame(rows)
                        dfs.append((safe_name, df))
                        sheets_written.append(sheet_name)
                    except Exception as sheet_err:
                        sheets_written.append(f"{safe_name}(error:{sheet_err})")
            elif "columns" in content and "rows" in content:
                df = pd.DataFrame(content["rows"], columns=content["columns"])
                sheet_name = str(content.get("sheet", "Sheet1"))[:31]
                dfs.append((sheet_name, df))
                sheets_written.append(sheet_name)
            else:
                safe_content = {str(k): v for k, v in content.items()}
                df = pd.DataFrame(safe_content)
                dfs.append(("Sheet1", df))
                sheets_written.append("Sheet1")

        elif isinstance(content, list):
            if content and isinstance(content[0], dict):
                df = pd.DataFrame(content)
            elif content and isinstance(content[0], list):
                df = pd.DataFrame(content[1:], columns=content[0])
            else:
                df = pd.DataFrame(content)
            dfs.append(("Sheet1", df))
            sheets_written.append("Sheet1")

        # Write and close in a tight scope to prevent handle leaks
        with pd.ExcelWriter(str(p), engine="openpyxl") as writer:
            for sheet_name, df in dfs:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Style header rows
            wb = writer.book
            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="2C3E50")
                    cell.alignment = Alignment(horizontal="center")
                for col_idx, col in enumerate(ws.columns, 1):
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value is not None),
                        default=8,
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(
                        max_len + 4, 40
                    )

        return {
            "status": "success",
            "path": str(p),
            "size": p.stat().st_size,
            "sheets_written": sheets_written,
        }
    except ImportError:
        return {"status": "error",
                "error": "pandas/openpyxl not installed. Run: pip install pandas openpyxl"}
    except Exception as e:
        return {"status": "error", "error": f"XLSX write failed: {type(e).__name__}: {e}"}
